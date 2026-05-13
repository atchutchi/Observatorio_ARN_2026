from pathlib import Path

from django.conf import settings
from django.core.management.base import BaseCommand, CommandError
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from apps.indicators.models import Indicator, IndicatorCategory


MONTH_LABELS = [
    'JANEIRO', 'FEVEREIRO', 'MARCO', 'ABRIL', 'MAIO', 'JUNHO',
    'JULHO', 'AGOSTO', 'SETEMBRO', 'OUTUBRO', 'NOVEMBRO', 'DEZEMBRO',
]

SHEET_LAYOUT = [
    ('estacoes_moveis', 'Estações móveis'),
    ('trafego_originado', 'Trafego_originado'),
    ('trafego_terminado', 'Trafego_Terminado'),
    ('trafego_roaming', 'Trafego_Roaming_Internacional'),
    ('internet_trafic', 'Internet_Trafic'),
    ('internet_fixo', 'Internet_Fixo'),
    ('lbi', 'LBI'),
    ('tarifario_voz', 'tarifario_voz'),
    ('receitas', 'RECEITAS'),
    ('empregos', 'Empregos'),
    ('investimento', 'Investimento'),
]

CUMULATIVE_CATEGORIES = {'receitas', 'investimento'}


class Command(BaseCommand):
    help = 'Gera um modelo Excel aceite pelo upload de indicadores da ARN'

    def add_arguments(self, parser):
        parser.add_argument(
            '--output',
            default=None,
            help='Caminho do ficheiro .xlsx a gerar.',
        )

    def handle(self, *args, **options):
        if not IndicatorCategory.objects.exists():
            raise CommandError('Sem categorias de indicadores. Execute seed_data primeiro.')

        output = options.get('output')
        if output:
            output_path = Path(output).expanduser()
            if not output_path.is_absolute():
                output_path = Path.cwd() / output_path
        else:
            output_path = (
                Path(settings.BASE_DIR).parent
                / 'frontend'
                / 'public'
                / 'templates'
                / 'modelo_upload_kpi_arn.xlsx'
            )

        output_path.parent.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)

        self._create_instruction_sheet(workbook)

        for category_code, sheet_name in SHEET_LAYOUT:
            category = IndicatorCategory.objects.filter(code=category_code).first()
            if not category:
                self.stdout.write(self.style.WARNING(
                    f'Categoria {category_code} não encontrada. Folha ignorada.',
                ))
                continue

            if category_code in CUMULATIVE_CATEGORIES:
                self._create_cumulative_sheet(workbook, category, sheet_name)
            else:
                self._create_monthly_sheet(workbook, category, sheet_name)

        workbook.save(output_path)
        self.stdout.write(self.style.SUCCESS(f'Modelo gerado: {output_path}'))

    def _create_instruction_sheet(self, workbook):
        sheet = workbook.create_sheet('Instruções')
        sheet.append(['Modelo de Upload de Indicadores ARN'])
        sheet.append([])
        rows = [
            ['Uso', 'Carregue este ficheiro em Entrada de Dados > Upload Excel.'],
            ['Operador', 'Seleccione o operador correcto antes do upload.'],
            ['Ano', 'Seleccione o ano civil dos dados.'],
            ['Trimestre', 'Seleccione Q1, Q2, Q3 ou Q4. A app importa apenas os meses desse trimestre.'],
            ['Folhas', 'Não altere os nomes das folhas.'],
            ['Códigos', 'Não altere a coluna Cod.'],
            ['Valores', 'Preencha apenas números. Use células vazias para dados não disponíveis.'],
            ['Receitas e Investimento', 'Preencha 3M, 6M, 9M ou 12M como valores acumulados.'],
            ['Questionários oficiais', 'Também pode carregar directamente ficheiros oficiais da Orange e Telecel/MTN.'],
        ]
        for row in rows:
            sheet.append(row)

        sheet['A1'].font = Font(bold=True, size=14, color='FFFFFF')
        sheet['A1'].fill = PatternFill('solid', fgColor='1F3A5F')
        sheet.merge_cells('A1:B1')
        sheet.column_dimensions['A'].width = 24
        sheet.column_dimensions['B'].width = 92
        for row in sheet.iter_rows(min_row=3, max_col=2):
            row[0].font = Font(bold=True)
            for cell in row:
                cell.alignment = Alignment(vertical='top', wrap_text=True)

    def _create_monthly_sheet(self, workbook, category, sheet_name):
        sheet = workbook.create_sheet(sheet_name[:31])
        sheet.append([category.name.upper()])
        sheet.append(['Cod.', 'INDICADOR', 'Unidade', 'Dados mensais'])
        sheet.append(['', '', '', 'Preencher os meses do trimestre seleccionado no upload.'])
        sheet.append(['Cod.', 'INDICADOR', 'Unidade', *MONTH_LABELS, 'Observações'])

        indicators = self._category_indicators(category)
        for indicator in indicators:
            sheet.append([
                indicator.code,
                self._display_name(indicator),
                indicator.get_unit_display(),
                *([''] * len(MONTH_LABELS)),
                'Calculado/opcional' if indicator.is_calculated else '',
            ])

        self._style_sheet(sheet, last_col=16, calculated_note_col=16)

    def _create_cumulative_sheet(self, workbook, category, sheet_name):
        sheet = workbook.create_sheet(sheet_name[:31])
        sheet.append([category.name.upper()])
        sheet.append(['N.º', 'INDICADOR', 'Unidade', 'Ano civil', 'Valores acumulados'])
        sheet.append(['', '', '', 'Ano', '3 MESES', '6 MESES', '9 MESES', '12 MESES', 'Observações'])
        sheet.append([
            'Cod.', 'INDICADOR', 'Unidade', 'Ano',
            '3 MESES', '6 MESES', '9 MESES', '12 MESES', 'Observações',
        ])

        indicators = self._category_indicators(category)
        for indicator in indicators:
            sheet.append([
                indicator.code,
                self._display_name(indicator),
                indicator.get_unit_display(),
                '',
                '',
                '',
                '',
                '',
                'Calculado/opcional' if indicator.is_calculated else '',
            ])

        self._style_sheet(sheet, last_col=9, calculated_note_col=9)

    @staticmethod
    def _category_indicators(category):
        return Indicator.objects.filter(category=category).order_by('order', 'code')

    @staticmethod
    def _display_name(indicator):
        indent = '  ' * max(indicator.level - 1, 0)
        return f'{indent}{indicator.name}'

    @staticmethod
    def _style_sheet(sheet, last_col, calculated_note_col):
        title_fill = PatternFill('solid', fgColor='1F3A5F')
        header_fill = PatternFill('solid', fgColor='D9EAF7')
        input_fill = PatternFill('solid', fgColor='FFF7D6')
        calculated_fill = PatternFill('solid', fgColor='F2F4F7')
        thin = Side(style='thin', color='D0D5DD')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
        sheet['A1'].font = Font(bold=True, size=13, color='FFFFFF')
        sheet['A1'].fill = title_fill
        sheet['A1'].alignment = Alignment(horizontal='center')

        for row_idx in [2, 3, 4]:
            for col_idx in range(1, last_col + 1):
                cell = sheet.cell(row_idx, col_idx)
                cell.font = Font(bold=True)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = border

        for row in sheet.iter_rows(min_row=5, max_col=last_col):
            is_calculated = row[calculated_note_col - 1].value == 'Calculado/opcional'
            for idx, cell in enumerate(row, start=1):
                cell.border = border
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                if is_calculated:
                    cell.fill = calculated_fill
                elif idx >= 4 and idx < calculated_note_col:
                    cell.fill = input_fill

        widths = {
            1: 16,
            2: 58,
            3: 18,
        }
        for col_idx in range(1, last_col + 1):
            width = widths.get(col_idx, 14)
            sheet.column_dimensions[get_column_letter(col_idx)].width = width

        sheet.freeze_panes = 'D5'
        sheet.auto_filter.ref = f'A4:{get_column_letter(last_col)}{sheet.max_row}'
