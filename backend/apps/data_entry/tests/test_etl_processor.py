from decimal import Decimal
import tempfile

from django.core.files import File
from django.test import TestCase, override_settings
from openpyxl import Workbook

from apps.data_entry.etl.processor import ExcelProcessor
from apps.data_entry.models import CumulativeData, DataEntry, FileUpload
from apps.indicators.models import (
    Indicator,
    IndicatorCategory,
    OperatorTypeIndicator,
    Period,
)
from apps.operators.models import Operator, OperatorType


class ExcelProcessorTest(TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.settings = override_settings(MEDIA_ROOT=self.tmp.name)
        self.settings.enable()

        self.operator_type = OperatorType.objects.create(
            code='terrestrial_full',
            name='Operador terrestre',
        )
        self.operator = Operator.objects.create(
            name='Telecel',
            legal_name='Spacetel',
            code='TELECEL',
            operator_type=self.operator_type,
        )
        for month in range(1, 13):
            Period.objects.create(
                year=2023,
                quarter=((month - 1) // 3) + 1,
                month=month,
                start_date=f'2023-{month:02d}-01',
                end_date=f'2023-{month:02d}-28',
            )

    def tearDown(self):
        self.settings.disable()
        self.tmp.cleanup()

    def test_imports_quarter_columns_from_real_questionnaire_layout(self):
        category = IndicatorCategory.objects.create(
            code='estacoes_moveis',
            name='Estações Móveis',
            order=1,
        )
        parent = Indicator.objects.create(
            category=category,
            code='1.1',
            name='Estações móveis em utilização efetiva',
            unit='number',
            level=1,
            order=1,
        )
        post_paid = Indicator.objects.create(
            category=category,
            code='1.1.1',
            name='Pós-Pago',
            unit='number',
            level=2,
            parent=parent,
            order=2,
        )
        pre_paid = Indicator.objects.create(
            category=category,
            code='1.1.2',
            name='Pré-Pago',
            unit='number',
            level=2,
            parent=parent,
            order=3,
        )
        OperatorTypeIndicator.objects.create(
            operator_type=self.operator_type,
            indicator=parent,
            is_applicable=True,
        )

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Estações móveis '
        sheet['A8'] = 'Cod.'
        sheet['B8'] = 'INDICADOR'
        sheet['D8'] = '3º TRIMESTRE 2023'
        sheet['G8'] = '4º TRIMESTRE 2023'
        for col, label in zip('DEFGHI', ['Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']):
            sheet[f'{col}10'] = label
        sheet['A12'] = '1.1'
        sheet['B12'] = 'Afectos a planos Pós-pagos'
        sheet['D12'] = 10
        sheet['E12'] = 11
        sheet['F12'] = 12
        sheet['G12'] = 99
        sheet['A13'] = '1.2'
        sheet['B13'] = 'Afectos a planos pré-pagos'
        sheet['D13'] = 100
        sheet['E13'] = 101
        sheet['F13'] = 102
        sheet['G13'] = 999

        upload = self._upload_workbook(workbook, quarter=3)

        self.assertTrue(ExcelProcessor(upload).process())
        self.assertEqual(
            DataEntry.objects.get(indicator=post_paid, period__month=7).value,
            Decimal('10.0000'),
        )
        self.assertEqual(
            DataEntry.objects.get(indicator=pre_paid, period__month=9).value,
            Decimal('102.0000'),
        )
        self.assertFalse(
            DataEntry.objects.filter(indicator=post_paid, period__month=10).exists(),
        )

    def test_monthly_cumulative_adds_previous_period_once_for_duplicate_targets(self):
        category = IndicatorCategory.objects.create(
            code='investimento',
            name='Investimento',
            order=1,
            is_cumulative=True,
        )
        target = Indicator.objects.create(
            category=category,
            code='1.1.1.1',
            name='Rede de acesso',
            unit='fcfa_millions',
            level=3,
            order=1,
        )
        OperatorTypeIndicator.objects.create(
            operator_type=self.operator_type,
            indicator=target,
            is_applicable=True,
        )
        CumulativeData.objects.create(
            indicator=target,
            operator=self.operator,
            year=2023,
            cumulative_type='6M',
            value=Decimal('100'),
        )

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Investimento'
        sheet['A2'] = 'N.º'
        sheet['B2'] = 'INDICADOR'
        sheet['D2'] = '3º TRIMESTRE 2023'
        sheet['D5'] = 'Julho'
        sheet['E5'] = 'Agosto'
        sheet['F5'] = 'Setembro'
        sheet['A8'] = '1.1'
        sheet['B8'] = 'Equipamentos de Telecomunicacoes Radios'
        sheet['D8'] = 10
        sheet['A9'] = '1.1'
        sheet['B9'] = 'Equipamentos de Telecomunicacoes Batarias Solares'
        sheet['D9'] = 1
        sheet['E9'] = 2
        sheet['F9'] = 3

        upload = self._upload_workbook(workbook, quarter=3)

        self.assertTrue(ExcelProcessor(upload).process())
        imported = CumulativeData.objects.get(
            indicator=target,
            operator=self.operator,
            year=2023,
            cumulative_type='9M',
        )
        self.assertEqual(imported.value, Decimal('116.0000'))

    def test_imports_internal_template_codes_directly(self):
        category = IndicatorCategory.objects.create(
            code='empregos',
            name='Empregos',
            order=1,
        )
        parent = Indicator.objects.create(
            category=category,
            code='1',
            name='Emprego Directo',
            unit='persons',
            level=1,
            order=1,
        )
        indicator = Indicator.objects.create(
            category=category,
            code='1.1',
            name='Nacionais',
            unit='persons',
            level=2,
            parent=parent,
            order=1,
        )
        OperatorTypeIndicator.objects.create(
            operator_type=self.operator_type,
            indicator=indicator,
            is_applicable=True,
        )

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Empregos'
        sheet['A4'] = 'Cod.'
        sheet['B4'] = 'INDICADOR'
        sheet['D4'] = 'JANEIRO'
        sheet['E4'] = 'FEVEREIRO'
        sheet['F4'] = 'MARCO'
        sheet['A6'] = '1.1'
        sheet['B6'] = 'Nacionais'
        sheet['D6'] = 93
        sheet['E6'] = 92
        sheet['F6'] = 92

        upload = self._upload_workbook(workbook, quarter=1)

        self.assertTrue(ExcelProcessor(upload).process())
        self.assertEqual(
            DataEntry.objects.get(indicator=indicator, period__month=1).value,
            Decimal('93.0000'),
        )
        self.assertEqual(
            DataEntry.objects.get(indicator=indicator, period__month=3).value,
            Decimal('92.0000'),
        )

    def _upload_workbook(self, workbook, quarter):
        path = f'{self.tmp.name}/questionario_q{quarter}.xlsx'
        workbook.save(path)
        upload = FileUpload(
            operator=self.operator,
            original_filename=f'questionario_q{quarter}.xlsx',
            file_type='questionnaire_telecel',
            year=2023,
            quarter=quarter,
        )
        with open(path, 'rb') as source:
            upload.file.save(f'questionario_q{quarter}.xlsx', File(source), save=False)
        upload.save()
        return upload
