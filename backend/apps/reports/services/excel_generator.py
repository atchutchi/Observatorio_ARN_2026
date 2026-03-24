import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from apps.indicators.models import IndicatorCategory, Indicator
from apps.operators.models import Operator
from apps.dashboards.services import DashboardService


OPERATOR_HEX = {
    'TELECEL': 'E30613',
    'ORANGE': 'FF6600',
    'STARLINK': '000000',
}


class ExcelReportGenerator:

    def __init__(self, year, quarter=None):
        self.year = year
        self.quarter = quarter
        self.wb = Workbook()
        self.wb.remove(self.wb.active)

    def generate(self):
        categories = IndicatorCategory.objects.all().order_by('order')
        for cat in categories:
            self._create_category_sheet(cat)

        self._create_summary_sheet()
        return self._to_bytes()

    def _create_category_sheet(self, category):
        ws = self.wb.create_sheet(title=category.name[:31])

        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='1B2A4A', end_color='1B2A4A', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )

        applicable_ops = DashboardService.get_applicable_operators(category.code)

        headers = ['Cód.', 'Indicador']
        for op in applicable_ops:
            headers.append(op.name)
        headers.append('Total')

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        indicators = Indicator.objects.filter(category=category).order_by('order')
        data = DashboardService.get_indicator_data(category.code, self.year, self.quarter)

        row = 2
        for ind in indicators:
            ws.cell(row=row, column=1, value=ind.code).border = thin_border

            name_cell = ws.cell(row=row, column=2, value=('  ' * ind.level) + ind.name)
            name_cell.border = thin_border
            if ind.level == 0:
                name_cell.font = Font(bold=True)

            total = 0
            for col_idx, op in enumerate(applicable_ops, 3):
                matching = [d for d in data if d['indicator_code'] == ind.code and d['operator_code'] == op.code]
                val = matching[0]['value'] if matching else None
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='right')
                cell.number_format = '#,##0'
                if val:
                    total += val

            total_cell = ws.cell(row=row, column=len(applicable_ops) + 3, value=total or None)
            total_cell.border = thin_border
            total_cell.alignment = Alignment(horizontal='right')
            total_cell.font = Font(bold=True)
            total_cell.number_format = '#,##0'

            row += 1

        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 45
        for i in range(3, len(headers) + 1):
            ws.column_dimensions[chr(64 + i)].width = 18

    def _create_summary_sheet(self):
        ws = self.wb.create_sheet(title='Resumo', index=0)

        header_font = Font(bold=True, color='FFFFFF', size=12)
        header_fill = PatternFill(start_color='1B2A4A', end_color='1B2A4A', fill_type='solid')

        ws.merge_cells('A1:D1')
        title_cell = ws.cell(row=1, column=1, value=f'Observatório Telecom GB — {self.year}')
        title_cell.font = Font(bold=True, size=14, color='1B2A4A')

        summary = DashboardService.get_summary(self.year, self.quarter)

        kpis = [
            ('Total Assinantes Móvel', summary['total_subscribers']),
            ('Volume de Negócios (FCFA)', summary['total_revenue']),
            ('Tráfego de Dados', summary['total_data_traffic']),
            ('Taxa de Penetração (%)', summary['penetration_rate']),
            ('Variação Assinantes (%)', summary['subscriber_change']),
            ('Operadores Activos', summary['active_operators']),
        ]

        for row, h in enumerate(['KPI', 'Valor'], 0):
            cell = ws.cell(row=3, column=row + 1, value=h)
            cell.font = header_font
            cell.fill = header_fill

        for i, (label, value) in enumerate(kpis, 4):
            ws.cell(row=i, column=1, value=label).font = Font(bold=True)
            val_cell = ws.cell(row=i, column=2, value=value)
            val_cell.number_format = '#,##0.0'

        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20

    def _to_bytes(self):
        buffer = io.BytesIO()
        self.wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
