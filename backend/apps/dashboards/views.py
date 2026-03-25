from datetime import datetime
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework import status

from apps.indicators.models import IndicatorCategory, Indicator
from apps.indicators.serializers import IndicatorCategorySerializer
from .services import DashboardService


class DashboardSummaryView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        year = int(request.query_params.get('year', datetime.now().year))
        quarter = request.query_params.get('quarter')
        quarter = int(quarter) if quarter else None

        data = DashboardService.get_summary(year, quarter)
        return Response(data)


class DashboardIndicatorView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, category_code):
        year = int(request.query_params.get('year', datetime.now().year))
        quarter = request.query_params.get('quarter')
        quarter = int(quarter) if quarter else None
        operator = request.query_params.get('operator')
        indicator_code = request.query_params.get('indicator_code')

        try:
            data = DashboardService.get_indicator_data(
                category_code, year, quarter, operator, indicator_code,
            )
        except IndicatorCategory.DoesNotExist:
            return Response(
                {'error': f'Categoria "{category_code}" não encontrada'},
                status=status.HTTP_404_NOT_FOUND,
            )

        category = IndicatorCategory.objects.get(code=category_code)
        indicators = Indicator.objects.filter(
            category=category, level=0,
        ).order_by('order').values('code', 'name', 'unit')

        return Response({
            'category': {
                'code': category.code,
                'name': category.name,
                'is_cumulative': category.is_cumulative,
            },
            'root_indicators': list(indicators),
            'data': data,
        })


class DashboardMarketShareView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        year = int(request.query_params.get('year', datetime.now().year))
        quarter = request.query_params.get('quarter')
        quarter = int(quarter) if quarter else None
        market = request.query_params.get('market', 'mobile')

        data = DashboardService.get_market_share(year, quarter, market)
        return Response({'market': market, 'year': year, 'data': data})


class DashboardTrendsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        category = request.query_params.get('category', 'estacoes_moveis')
        indicator = request.query_params.get('indicator')
        start_year = int(request.query_params.get('start_year', 2018))
        end_year = int(request.query_params.get('end_year', datetime.now().year))

        data = DashboardService.get_trends(category, indicator, start_year, end_year)

        operators = DashboardService.get_applicable_operators(category)
        operator_info = [
            {'code': op.code, 'name': op.name, 'color': op.brand_color}
            for op in operators
        ]

        return Response({
            'category': category,
            'indicator': indicator,
            'operators': operator_info,
            'data': data,
        })


class DashboardComparativeView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        year = int(request.query_params.get('year', datetime.now().year))
        categories = request.query_params.getlist('categories')
        if not categories:
            categories = list(
                IndicatorCategory.objects.values_list('code', flat=True),
            )

        category_to_market = {
            'estacoes_moveis': 'mobile',
            'trafego_originado': 'voice',
            'trafego_terminado': 'voice',
            'trafego_roaming': 'mobile',
            'internet_fixo': 'fixed_internet',
            'internet_trafic': 'data',
            'receitas': 'revenue',
            'empregos': 'employment',
            'investimento': 'revenue',
            'lbi': 'data',
            'tarifario_voz': 'voice',
        }

        results = {}
        for cat_code in categories:
            try:
                growth = DashboardService.get_growth_rates(cat_code, year)
                market_key = category_to_market.get(cat_code, 'mobile')
                market = DashboardService.get_market_share(year, market=market_key)
                results[cat_code] = {
                    'growth': growth,
                    'market_share': market,
                }
            except IndicatorCategory.DoesNotExist:
                continue

        return Response({'year': year, 'categories': results})


class DashboardCAGRView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        category = request.query_params.get('category', 'estacoes_moveis')
        start_year = int(request.query_params.get('start_year', 2018))
        end_year = int(request.query_params.get('end_year', datetime.now().year))

        data = DashboardService.get_cagr(category, start_year, end_year)
        return Response({
            'category': category,
            'start_year': start_year,
            'end_year': end_year,
            'data': data,
        })


class DashboardHHIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        year = int(request.query_params.get('year', datetime.now().year))
        market = request.query_params.get('market', 'mobile')

        data = DashboardService.get_hhi(year, market)
        return Response(data)


class DashboardExportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        category = request.query_params.get('category')
        year = int(request.query_params.get('year', datetime.now().year))
        fmt = request.query_params.get('format', 'json')

        if not category:
            return Response(
                {'error': 'Parâmetro "category" obrigatório'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        data = DashboardService.get_indicator_data(category, year)

        if fmt == 'json':
            return Response(data)

        from django.http import HttpResponse
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active

        cat = IndicatorCategory.objects.get(code=category)
        ws.title = cat.name[:31]

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1B2A4A', end_color='1B2A4A', fill_type='solid')

        headers = ['Indicador', 'Código', 'Operador', 'Período', 'Valor', 'Unidade']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        for row_idx, item in enumerate(data, 2):
            ws.cell(row=row_idx, column=1, value=item['indicator_name'])
            ws.cell(row=row_idx, column=2, value=item['indicator_code'])
            ws.cell(row=row_idx, column=3, value=item['operator_name'])
            ws.cell(row=row_idx, column=4, value=item['period'])
            ws.cell(row=row_idx, column=5, value=item['value'])
            ws.cell(row=row_idx, column=6, value=item['unit'])

        for col in range(1, 7):
            ws.column_dimensions[chr(64 + col)].width = 20

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{category}_{year}.xlsx"'
        wb.save(response)
        return response
